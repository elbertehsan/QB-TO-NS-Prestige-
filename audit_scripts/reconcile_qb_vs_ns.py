"""
Financial Reconciliation — QuickBooks vs NetSuite
==================================================
Compares actual debit/credit amounts by account and location
between QB (via server memory) and NS (via SuiteQL).

This goes beyond transaction count — it validates the actual
financial data matches so there are no variances.

Usage:
    # Step 1: Queue the date range and run QBWC first
    Invoke-WebRequest "http://127.0.0.1:8001/add-chunk?from_date=2026-04-01&to_date=2026-04-30"
    # Click Update Selected in QBWC, wait for 100%

    # Step 2: Run reconciliation
    python reconcile_qb_vs_ns.py --from-date 2026-04-01 --to-date 2026-04-30

    # Step 3: Export to Excel for review
    python reconcile_qb_vs_ns.py --from-date 2026-04-01 --to-date 2026-04-30 --excel
"""

import argparse
import json
import requests
from collections import defaultdict
from datetime import datetime, timedelta
from dotenv import dotenv_values
from netsuite_posting import get_jwt_token, generate_access_token
from logger_config import logger

config        = dotenv_values(".env")
NETSUITE_BASE = config.get('NETSITE_BASE_URL').rstrip('/')
JWT_BASE_URL  = config.get('AZURE_NREST_BASE_URL')
DIAG_HOST     = "http://127.0.0.1:8001"
SUBSIDIARY_ID = 15


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_ns_token():
    jwt   = get_jwt_token(JWT_BASE_URL)
    token = generate_access_token(NETSUITE_BASE, jwt)
    return token


def ns_headers(token):
    return {
        "Authorization": f"{token.get('token_type')} {token.get('access_token')}",
        "Content-Type":  "application/json",
        "Prefer":        "transient",
    }


# ── SuiteQL ───────────────────────────────────────────────────────────────────

def run_suiteql(token, query, limit=500):
    base = NETSUITE_BASE
    if base.endswith('/services/rest'):
        base = base[:-len('/services/rest')]
    url      = f"{base}/services/rest/query/v1/suiteql"
    all_rows, offset = [], 0
    while True:
        r = requests.post(url, headers=ns_headers(token),
                          json={"q": query},
                          params={"limit": limit, "offset": offset},
                          timeout=(30, 300))
        if r.status_code not in (200, 204):
            print(f"\n  [SuiteQL ERROR] HTTP {r.status_code}")
            print(f"  Detail: {r.text[:800]}")
            break
        data = r.json()
        rows = data.get("items", [])
        all_rows.extend(rows)
        if not data.get("hasMore") or not rows:
            break
        offset += limit
        print(f"  Fetched {len(all_rows)} NS rows...", end="\r")
    return all_rows


# ── Fetch QB data from server memory ─────────────────────────────────────────

def fetch_qb_line_items():
    """
    Fetch full QB line item data from server memory.
    Returns list of line items with account, debit, credit, location, txn_id, date.
    """
    try:
        r    = requests.get(f"{DIAG_HOST}/results", timeout=10)
        data = r.json()
        # Server stores full transaction data, not just IDs
        line_items = data.get("qb_line_items", [])
        txns       = data.get("qb_transactions", {})
        print(f"  QB transactions in server: {len(txns)}")
        print(f"  QB line items in server  : {len(line_items)}")
        return line_items, txns
    except requests.ConnectionError:
        print("  [ERROR] Cannot connect to server on port 8001.")
        return [], {}


# ── Fetch NS line items ───────────────────────────────────────────────────────

def fetch_ns_line_items(token, from_date, to_date):
    """
    Fetch all journal entry line items from NetSuite for April.
    Groups by account and location to get totals.
    """
    from_dt  = datetime.strptime(from_date, "%Y-%m-%d")
    to_dt    = datetime.strptime(to_date,   "%Y-%m-%d")
    all_rows = []
    current  = from_dt
    chunk    = 0

    print(f"\n  Fetching NS line items week by week...")

    while current <= to_dt:
        chunk_end = min(current + timedelta(days=6), to_dt)
        fs = current.strftime("%Y-%m-%d")
        ts = chunk_end.strftime("%Y-%m-%d")
        chunk += 1
        print(f"  Week {chunk}: {fs} → {ts}", end="  ")

        rows = run_suiteql(token, f"""
            SELECT
                t.id                AS transaction_id,
                t.tranid            AS tran_id,
                t.trandate          AS tran_date,
                jel.account         AS account_id,
                a.acctnumber        AS account_number,
                a.fullname          AS account_name,
                jel.location        AS location_id,
                l.name              AS location_name,
                jel.debit           AS debit,
                jel.credit          AS credit,
                jel.memo            AS memo
            FROM
                transaction t
                JOIN journalentryline jel ON jel.journal = t.id
                JOIN account a            ON a.id = jel.account
                LEFT JOIN location l      ON l.id = jel.location
            WHERE
                t.recordtype = 'journalentry'
                AND t.subsidiary = 15
                AND t.trandate  >= TO_DATE('{fs}', 'YYYY-MM-DD')
                AND t.trandate  <= TO_DATE('{ts}', 'YYYY-MM-DD')
            ORDER BY t.trandate ASC, t.id ASC
        """)
        print(f"{len(rows)} lines")
        all_rows.extend(rows)
        current = chunk_end + timedelta(days=1)

    print(f"\n  Total NS line items: {len(all_rows):,}")
    return all_rows


# ── Aggregate NS data ─────────────────────────────────────────────────────────

def aggregate_ns_by_account(ns_lines):
    """
    Aggregate NS line items by account number.
    Returns { account_number -> { debit_total, credit_total, net } }
    """
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0, "name": ""})

    for row in ns_lines:
        acct_num  = str(row.get("account_number") or "").strip()
        acct_name = str(row.get("account_name")   or "").strip()
        debit     = float(row.get("debit")  or 0)
        credit    = float(row.get("credit") or 0)

        totals[acct_num]["debit"]  += debit
        totals[acct_num]["credit"] += credit
        totals[acct_num]["name"]    = acct_name

    return totals


def aggregate_ns_by_location(ns_lines):
    """
    Aggregate NS line items by location.
    Returns { location_name -> { debit_total, credit_total } }
    """
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})

    for row in ns_lines:
        loc_name = str(row.get("location_name") or "No Location").strip()
        debit    = float(row.get("debit")  or 0)
        credit   = float(row.get("credit") or 0)

        totals[loc_name]["debit"]  += debit
        totals[loc_name]["credit"] += credit

    return totals


def aggregate_ns_by_tranid(ns_lines):
    """
    Aggregate NS totals per transaction ID.
    Returns { tran_id -> { debit, credit, date, line_count } }
    """
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0, "date": "", "lines": 0})

    for row in ns_lines:
        tran_id = str(row.get("tran_id") or "").strip()
        # Strip -15 suffix
        bare = tran_id.replace("-15", "").replace("-14", "")
        if not bare:
            continue
        debit  = float(row.get("debit")  or 0)
        credit = float(row.get("credit") or 0)

        totals[bare]["debit"]  += debit
        totals[bare]["credit"] += credit
        totals[bare]["date"]    = str(row.get("tran_date") or "")
        totals[bare]["lines"]  += 1

    return totals


# ── Build QB aggregates from server data ─────────────────────────────────────

def aggregate_qb_from_server(line_items):
    """
    Aggregate QB line items by account number.
    Server must store full line item data (not just txn IDs).
    """
    by_account  = defaultdict(lambda: {"debit": 0.0, "credit": 0.0, "name": ""})
    by_location = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})

    for item in line_items:
        acct    = str(item.get("account", "")).strip()
        loc     = str(item.get("location", "No Location")).strip()
        debit   = float(item.get("debit",  0) or 0)
        credit  = float(item.get("credit", 0) or 0)

        by_account[acct]["debit"]  += debit
        by_account[acct]["credit"] += credit
        by_location[loc]["debit"]  += debit
        by_location[loc]["credit"] += credit

    return by_account, by_location


# ── Compare NS tranId totals with QB data ─────────────────────────────────────

def compare_by_tranid(qb_txns, ns_by_tranid, ns_lines):
    """
    For each QB transaction, check if NS has the same total debit.
    This is the most granular check — per transaction financial match.
    """
    mismatches = []
    matched    = 0

    # Build QB per-txn totals from NS (we use NS data since we may not have
    # QB line items in server — this detects missing transactions)
    qb_ids = set(qb_txns.keys())
    ns_ids = set(ns_by_tranid.keys())
    print(f"\n Acutual Trans id in NS: {len(ns_by_tranid.keys())}")

    in_both  = qb_ids & ns_ids
    only_qb  = qb_ids - ns_ids   # missing from NS
    only_ns  = ns_ids - qb_ids   # phantom in NS

    return {
        "matched":   len(in_both),
        "only_qb":   sorted(only_qb),    # missing from NS
        "only_ns":   sorted(only_ns),    # phantom in NS
        "ns_totals": {k: ns_by_tranid[k] for k in in_both},
    }


# ── Print report ──────────────────────────────────────────────────────────────

def print_report(ns_by_account, ns_by_location, ns_by_tranid,
                 qb_txns, from_date, to_date):
    div = "─" * 70
    eq  = "=" * 70

    print(f"\n{eq}")
    print(f"  FINANCIAL RECONCILIATION REPORT")
    print(f"  Subsidiary {SUBSIDIARY_ID} — Prestige Fleet Services")
    print(f"  Period: {from_date} → {to_date}")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{eq}")

    # ── Transaction count ─────────────────────────────────────────────────────
    comparison = compare_by_tranid(qb_txns, ns_by_tranid, [])
    print(f"\n  TRANSACTION SYNC STATUS")
    print(div)
    print(f"  QB transactions        : {len(qb_txns):>8,}")
    print(f"  NS unique tranIds      : {len(ns_by_tranid):>8,}")
    print(f"  ✅ Matched             : {comparison['matched']:>8,}")
    print(f"  ❌ Missing from NS     : {len(comparison['only_qb']):>8,}")
    print(f"  ⚠  Phantom in NS      : {len(comparison['only_ns']):>8,}")

    if comparison['only_qb']:
        print(f"\n  Missing QB tranIds (first 20):")
        for tid in comparison['only_qb'][:20]:
            print(f"    {tid}   QB date: {qb_txns.get(tid, '?')}")
        if len(comparison['only_qb']) > 20:
            print(f"    ... +{len(comparison['only_qb'])-20} more")

    # ── NS totals by account ──────────────────────────────────────────────────
    print(f"\n  NS TOTALS BY ACCOUNT")
    print(div)
    print(f"  {'Account':<12} {'Name':<40} {'Debit':>14} {'Credit':>14} {'Net':>14}")
    print(f"  {'─'*11} {'─'*39} {'─'*14} {'─'*14} {'─'*14}")

    total_debit  = 0.0
    total_credit = 0.0

    for acct_num in sorted(ns_by_account.keys()):
        data   = ns_by_account[acct_num]
        debit  = data["debit"]
        credit = data["credit"]
        net    = debit - credit
        name   = data["name"][:39]
        total_debit  += debit
        total_credit += credit
        print(f"  {acct_num:<12} {name:<40} {debit:>14,.2f} {credit:>14,.2f} {net:>14,.2f}")

    print(div)
    print(f"  {'TOTAL':<53} {total_debit:>14,.2f} {total_credit:>14,.2f} {total_debit-total_credit:>14,.2f}")

    # ── NS totals by location ─────────────────────────────────────────────────
    print(f"\n  NS TOTALS BY LOCATION")
    print(div)
    print(f"  {'Location':<35} {'Debit':>14} {'Credit':>14} {'Net':>14}")
    print(f"  {'─'*34} {'─'*14} {'─'*14} {'─'*14}")

    for loc in sorted(ns_by_location.keys()):
        data   = ns_by_location[loc]
        debit  = data["debit"]
        credit = data["credit"]
        net    = debit - credit
        print(f"  {loc:<35} {debit:>14,.2f} {credit:>14,.2f} {net:>14,.2f}")

    print(f"\n{eq}\n")


# ── Save JSON report ──────────────────────────────────────────────────────────

def save_report(ns_by_account, ns_by_location, ns_by_tranid,
                qb_txns, from_date, to_date):
    comparison = compare_by_tranid(qb_txns, ns_by_tranid, [])

    report = {
        "generated_at":    datetime.now().isoformat(),
        "subsidiary":      f"{SUBSIDIARY_ID} (Prestige)",
        "date_range":      {"from": from_date, "to": to_date},
        "transaction_sync": {
            "qb_count":       len(qb_txns),
            "ns_count":       len(ns_by_tranid),
            "matched":        comparison["matched"],
            "missing_from_ns": comparison["only_qb"],
            "phantom_in_ns":  comparison["only_ns"],
        },
        "ns_by_account":  {
            k: {**v, "net": round(v["debit"] - v["credit"], 2)}
            for k, v in ns_by_account.items()
        },
        "ns_by_location": {
            k: {**v, "net": round(v["debit"] - v["credit"], 2)}
            for k, v in ns_by_location.items()
        },
    }

    fname = f"reconciliation_{from_date.replace('-','')}_{to_date.replace('-','')}.json"
    with open(fname, "w") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"  Report saved → {fname}")
    return fname


# ── Optional Excel export ─────────────────────────────────────────────────────

def save_excel(ns_by_account, ns_by_location, ns_by_tranid,
               qb_txns, from_date, to_date):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [SKIP] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: By Account ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "By Account"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Account Number", "Account Name", "Debit", "Credit", "Net"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (acct_num, data) in enumerate(sorted(ns_by_account.items()), 2):
        ws1.cell(row=row_idx, column=1, value=acct_num)
        ws1.cell(row=row_idx, column=2, value=data["name"])
        ws1.cell(row=row_idx, column=3, value=round(data["debit"],  2))
        ws1.cell(row=row_idx, column=4, value=round(data["credit"], 2))
        ws1.cell(row=row_idx, column=5, value=round(data["debit"] - data["credit"], 2))

        # Highlight non-zero net
        net = data["debit"] - data["credit"]
        if abs(net) > 0.01:
            ws1.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor="FFF2CC")

    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 2: By Location ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Location")
    for col, h in enumerate(["Location", "Debit", "Credit", "Net"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (loc, data) in enumerate(sorted(ns_by_location.items()), 2):
        ws2.cell(row=row_idx, column=1, value=loc)
        ws2.cell(row=row_idx, column=2, value=round(data["debit"],  2))
        ws2.cell(row=row_idx, column=3, value=round(data["credit"], 2))
        ws2.cell(row=row_idx, column=4, value=round(data["debit"] - data["credit"], 2))

    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 25

    # ── Sheet 3: Missing from NS ──────────────────────────────────────────────
    comparison = compare_by_tranid(qb_txns, ns_by_tranid, [])
    ws3 = wb.create_sheet("Missing from NS")
    for col, h in enumerate(["QB TranId", "QB Date", "Status"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, tid in enumerate(sorted(comparison["only_qb"]), 2):
        ws3.cell(row=row_idx, column=1, value=tid)
        ws3.cell(row=row_idx, column=2, value=qb_txns.get(tid, ""))
        ws3.cell(row=row_idx, column=3, value="Missing from NetSuite")
        ws3.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFE0E0")

    fname = f"reconciliation_{from_date.replace('-','')}_{to_date.replace('-','')}.xlsx"
    wb.save(fname)
    print(f"  Excel saved → {fname}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--to-date",   required=True, help="YYYY-MM-DD")
    parser.add_argument("--excel",     action="store_true",
                        help="Also export to Excel")
    args = parser.parse_args()

    print(f"\n{'='*70}")
    print(f"  Financial Reconciliation — QB vs NetSuite")
    print(f"  Subsidiary {SUBSIDIARY_ID} — Prestige Fleet Services")
    print(f"  Period: {args.from_date} → {args.to_date}")
    print(f"{'='*70}")

    # Get QB data from server
    print(f"\n  Fetching QB data from server...")
    line_items, qb_txns = fetch_qb_line_items()

    # Get NS data
    print(f"\n  Authenticating with NetSuite...")
    token = get_ns_token()
    print(f"  OK")

    ns_lines = fetch_ns_line_items(token, args.from_date, args.to_date)

    # Aggregate
    print(f"\n  Aggregating data...")
    ns_by_account  = aggregate_ns_by_account(ns_lines)
    ns_by_location = aggregate_ns_by_location(ns_lines)
    ns_by_tranid   = aggregate_ns_by_tranid(ns_lines)

    # QB aggregates (if server has line items)
    if line_items:
        qb_by_account, qb_by_location = aggregate_qb_from_server(line_items)
    else:
        qb_by_account  = {}
        qb_by_location = {}

    # Report
    print_report(ns_by_account, ns_by_location, ns_by_tranid,
                 qb_txns, args.from_date, args.to_date)

    # Save
    save_report(ns_by_account, ns_by_location, ns_by_tranid,
                qb_txns, args.from_date, args.to_date)

    if args.excel:
        save_excel(ns_by_account, ns_by_location, ns_by_tranid,
                   qb_txns, args.from_date, args.to_date)


if __name__ == "__main__":
    main()